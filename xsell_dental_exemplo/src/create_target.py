import logging
from pathlib import Path
import pandas as pd
import dask.dataframe as dd
from dateutil.relativedelta import relativedelta
from datetime import datetime

from ..utils import universe_creation as uc
from ..utils.config_reader import read_config

def create_target1_geral(lob, month):
    # Function to return the unique IDs by month

    return [id_column].unique().tolist()

# Definition of target function
def target_function(lob,month):

    return TARGET1_mes_menos2, TARGET0_Total_mes_menos2

def report_target0_1 (df_target0, df_target1):
    df_target0.loc[:,'TARGET'] = 0
    df_target1.loc[:,'TARGET'] = 1

    cols = ["ANO_MES", "ID_CLIENTE_ANON", "TARGET"]
    df_target1[cols] = df_target1[cols].astype(int)
    df_target0[cols] = df_target0[cols].astype(int)
    
    Universo = pd.concat([df_target0, df_target1], ignore_index=True)
    report_df = Universo.groupby('ANO_MES').agg({'TARGET':['count','sum','mean']})

    report_df.columns = ['# '+id_column, '# TARGET 1','% TARGET 1']
    report_df['# TARGET 0'] = report_df['# '+id_column] - report_df['# TARGET 1']
    report_df['% TARGET 1'] = report_df['% TARGET 1']*100
    
    return report_df[['# '+id_column, '# TARGET 0', '# TARGET 1','% TARGET 1']]

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

# Function to generate the plot for a given datadf
def generate_plot(df, sheet_name):
    
    df = df.reset_index()
    df['ANO_MES'] = df['ANO_MES'].astype(int).round(0)
    
    fig, ax1 = plt.subplots(figsize=(12, 6))

    # Assuming the df has the necessary columns: ANO_MES, TARGET 0, TARGET 1, % TARGET
    ax1.bar(df['ANO_MES'], df['# TARGET 0'], label='', color='lightgray') #Customize label
    ax1.bar(df['ANO_MES'], df['# TARGET 1'], bottom=df['# TARGET 0'], label='', color='red') #Customize label

    # Primary axis labels
    ax1.set_xlabel('Ano-Mês')
    ax1.set_ylabel('') #Customize label
    ax1.set_ylim([0, (df['# TARGET 0']+df['# TARGET 1']).max()*1.05])
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{int(x/1000)}k'))
    ax1.legend(loc='upper left')

    # Secondary axis for percentage
    ax2 = ax1.twinx()
    ax2.plot(df['ANO_MES'], df['% TARGET 1'], label='', color='black', marker='o', linestyle='-', linewidth=2) #Customize label
    ax2.set_ylabel('') #Customize label
    ax2.set_ylim([0, df['% TARGET 1'].max()*1.25])
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:.1f}%'))

    # Add annotations to the plot
    for i, txt in enumerate(df['% TARGET 1']):
        ax2.annotate(f'{txt:.2f}%', (df['ANO_MES'][i], df['% TARGET 1'][i]), textcoords="offset points", xytext=(0,5), ha='center')

    ax1.grid(True, axis='y', linestyle='--', alpha=0.6)
    ax2.grid(False)


    # Rotate x-axis labels and add title
    plt.xticks(rotation=45)
    plt.title(f'Número de ... por mês ({sheet_name})') #Customize label

    # Save plot as an image to be added to Excel later
    image_path = f'/tmp/{sheet_name}_plot.png'
    plt.tight_layout()
    plt.savefig(image_path)
    plt.close()  # Close the figure to avoid display overhead
    return image_path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def save_report_with_plots(report_dict, excel_path):
    file_name = 'Target Analysis.xlsx'
    with pd.ExcelWriter(excel_path / file_name, engine='openpyxl') as writer:
        for sheet, frame in report_dict.items():
            # Write DataFrame to Excel sheet
            frame.to_excel(writer, sheet_name=sheet, index=True)
        

    # Load the workbook with openpyxl for image manipulation
    wb = load_workbook(excel_path / file_name)

    for sheet, frame in report_dict.items():
        if '% TARGET 1' in frame.columns:
            # Generate the plot for each DataFrame
            plot_image_path = generate_plot(frame, sheet)

            # Get the worksheet for the current DataFrame
            ws = wb[sheet]

            # Try to embed the plot image in the same sheet
            try:
                img = Image(plot_image_path)
                ws.add_image(img, 'H2')  # Adjust the position of the image in the sheet
            except Exception as e:
                print(f"Failed to insert image in {sheet}: {e}")
                # Fallback to adding the image in a new sheet if it doesn't fit
                graphic_sheet_name = sheet + '_graphic'
                wb.create_sheet(graphic_sheet_name)
                ws_graphic = wb[graphic_sheet_name]
                ws_graphic.add_image(img, 'B2')  # Insert in a new sheet with a similar name

    # Save the workbook with all modifications
    wb.save(excel_path / file_name)


if __name__ == "__main__":


    PROJECT_PATH = Path(__file__).parents[1]
    DATA_PATH = uc.find_data_path(PROJECT_PATH) / "parquet"
    OUTPUT_PATH = PROJECT_PATH / "data"
    CONFIG_PATH = PROJECT_PATH / "configs"
    REPORT_AUX_PATH = PROJECT_PATH / "report" / "Auxiliary Reports"

    # Create logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # Create file handler
    file_handler = logging.FileHandler(PROJECT_PATH / "src" / "logfile.log")
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    # Create console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)


    config = read_config(CONFIG_PATH / "create_universe_config.yml")
    lob = config.lob
    client_type = config.client_type
    start_month = config.start_month
    end_month = config.end_month
    id_column = config.id_column

    # Import elIgibles dataset
    ELEGIVEIS = pd.read_csv(OUTPUT_PATH / f'eligibles_{start_month.strftime("%Y%m")}_{end_month.strftime("%Y%m")}.csv', sep=",", encoding='latin1')

    logging.info(ELEGIVEIS.shape)
    logging.info(ELEGIVEIS['ANO_MES'].value_counts().sort_index())
    
    save_reports = input("Do you want to save data analysis reports?[y/n]")
    if save_reports.lower() == 'y' or save_reports == '':
        save_reports = True
        report_dict = {}
        report_dict["Sem Target 1 Geral"] = pd.DataFrame()
        report_dict["Com Target 1 Geral"] = pd.DataFrame()
    else:
        save_reports = False

    if save_reports:
        report_df = ELEGIVEIS.groupby('ANO_MES')[[id_column]].nunique().rename(columns={id_column:'# '+id_column})
        report_dict['Eligibles'] = report_df

    # Aplication of target function

    TARGET1=pd.DataFrame()
    TARGET0=pd.DataFrame()
    t0_id = []
    t1_geral = set()
    use_target1_geral = config.target1_universe
    
    if use_target1_geral == "y":
    # Definition of list of months for target1 geral
        t1_geral_report = {}
        months = uc.generate_month(start_month - relativedelta(months=2), end_month + relativedelta(months=+3))
        for month in months:
            t1_geral_report[month] = create_target1_geral(lob, month)
            t1_geral = t1_geral.union(t1_geral_report[month])

        if save_reports:
            t1_geral_report = pd.DataFrame([(key, value) for key, values in t1_geral_report.items() for value in values]
                                           , columns=['ANO_MES', 'ID_APOLICE'])
            t1_geral_report_agg = t1_geral_report.groupby('ANO_MES')[['ID_APOLICE']].nunique().rename(columns={'ID_APOLICE':'# ID_APOLICE'})
            t1_geral_report_agg.loc['TOTAL UNIQUE','# ID_APOLICE'] = t1_geral_report['ID_APOLICE'].nunique()
            report_dict['Target1 Geral'] = t1_geral_report_agg      
    
    
    # Definition of list of months
    months = uc.generate_month(start_month, end_month)

    for month in months:
        logging.info(f"Creating target for month: {str(month)}")

        # Apply target function
        TARGET1_mes_menos2, TARGET0_mes_menos2 = target_function(lob, month)
        
        if save_reports:
            report_dict["Com Target 1 Geral"] = pd.concat([report_dict["Com Target 1 Geral"], report_target0_1(TARGET0_mes_menos2, TARGET1_mes_menos2)])
      

        TARGET0_SemT1G_mes_menos2 = TARGET0_mes_menos2[(~TARGET0_mes_menos2[id_column].isin(t1_geral))]
        
        if save_reports:
            report_dict["Sem Target 1 Geral"] = pd.concat([report_dict["Sem Target 1 Geral"], report_target0_1(TARGET0_SemT1G_mes_menos2, TARGET1_mes_menos2)])

        # Calculate the number of target0 by month, and number of months
        NTarget0 = TARGET0_SemT1G_mes_menos2.shape[0]
        Nmeses = len(months)

        # Filter and sample the data
        TARGET0_mes_menos2 = uc.filter_and_sample(TARGET0_SemT1G_mes_menos2, id_column, t0_id, int(NTarget0 / Nmeses))

        # Collect unique IDs for further filtering
        t0_id += TARGET0_mes_menos2[id_column].unique().tolist()

        # Create Target0 universe
        TARGET0 = pd.concat([TARGET0, TARGET0_mes_menos2], ignore_index=True)   
                
        # Create Target1 universe
        TARGET1 = pd.concat([TARGET1, TARGET1_mes_menos2], ignore_index=True)


    del TARGET1_mes_menos2
    del TARGET0_mes_menos2
    
    if save_reports:
        report_dict["Com amostragem"] = report_target0_1(TARGET0, TARGET1)

    TARGET0['TARGET']=0
    TARGET1['TARGET']=1
    Universo = pd.concat([TARGET0, TARGET1], ignore_index=True)
    logging.info(Universo['ANO_MES'].value_counts().sort_index())
    
    if save_reports:
        save_report_with_plots(report_dict, REPORT_AUX_PATH)
                             
    Universo.to_csv(OUTPUT_PATH / f'universe_{start_month.strftime("%Y%m")}_{end_month.strftime("%Y%m")}.csv', index=False, encoding="latin1")
